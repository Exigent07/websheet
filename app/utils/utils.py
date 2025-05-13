# HyDE default layout
# Wallbash generated layout

# Local Variables
$fontFamily = JetBrainsMono Nerd Font
$fn_greet=echo "Good $(date +%H | awk '{if ($1 < 12) print "Morning"; else if ($1 < 18) print "Afternoon"; else print "Evening"}'), ${USER}"

# GENERAL
general {
    no_fade_in = false
    grace = 0
    disable_loading_bar = true
}

# BACKGROUND
background {
    monitor =
    path = $BACKGROUND_PATH
    blur_passes = 2
}

# TIME
label {
  monitor =
  text = $TIME
  color = $wallbash_txt1_rgba
  font_size = 90
  font_family = $fontFamily
  position = -30, 0
  halign = right
  valign = top
}

# DATE
label {
  monitor =
  text = cmd[update:43200000] date +"%A, %d %B %Y"
  color = $wallbash_txt2_rgba
  font_size = 25
  font_family = $fontFamily
  position = -30, -150
  halign = right
  valign = top
}

# USER AVATAR
image {
    monitor =
    path = $MPRIS_IMAGE
    size = 150 # lesser side if not 1:1 ratio
    reload_time = 0 # seconds between reloading, 0 to reload with SIGUSR2
    reload_cmd = # command to get new path. if empty, old path will be used. don't run "follow" commands like tail -F
    position = 0, 0
    halign = center
    valign = center
}

# INPUT FIELD
input-field {
    monitor =
    size = 200, 50 #!
    outline_thickness = 3
    dots_size = 0.33 # Scale of input-field height, 0.2 - 0.8
    dots_spacing = 0.15 # Scale of dots' absolute size, 0.0 - 1.0
    dots_center = true
    dots_rounding = -1 # -1 default circle, -2 follow input-field rounding
    outer_color = $wallbash_pry4_rgba
    inner_color = $wallbash_pry2_rgba
    font_color = $wallbash_3xa9_rgba
    fade_on_empty = true
    fade_timeout = 1000 # Milliseconds before fade_on_empty is triggered.
    placeholder_text = <i>Input Password...</i> # Text rendered in the input box when it's empty.
    hide_input = false
    rounding = -1 # -1 means complete rounding (circle/oval)
    check_color = $wallbash_pry4_rgba
    fail_color = rgba(FF0000FF) # if authentication failed, changes outer_color and fail message color
    fail_text = <i>$FAIL <b>($ATTEMPTS)</b></i> # can be set to empty
    fail_transition = 300 # transition time in ms between normal outer_color and fail_color
    capslock_color = -1
    numlock_color = -1
    bothlock_color = -1 # when both locks are active. -1 means don't change outer color (same for above)
    invert_numlock = false # change color if numlock is off
    swap_font_color = true # see below
    position = 0, 80
    halign = center
    valign = bottom
}


# USER Greeting
label {
    monitor =
    text = cmd[update:60000] $fn_greet
    color = $text
    font_size = 20
    font_family = $fontFamily Bold
    position = 0, -190
    halign = center
    valign = center
}

# Mpris and SPLASH
label {
    monitor =
    text = cmd[update:1000] $SPLASH_CMD # Outputs the song title when mpris is available, otherwise, it will output the splash command.
    color = $wallbash_txt2_rgba
    font_family = $fontFamily
    font_size = 15
    position = 0, 0
    halign = center
    valign = bottom
}



# Battery Status if present
label {
    monitor =
    text = cmd[update:5000] $BATTERY_ICON
    color = $wallbash_4xa9_rgba
    font_size = 20
    font_family = JetBrainsMono Nerd Font
    position = -1%, 1%
    halign = right
    valign = bottom
}

# Current Keyboard Layout 
label {
    monitor =
    text = cmd[update:1000] $KEYBOARD_LAYOUT
    color = $wallbash_4xa9_rgba
    font_size = 20
    font_family = $fontFamily
    position = -2%, 1%
    halign = right
    valign = bottom
}
import csv
from pathlib import Path
from typing import Literal, Dict, List
import openpyxl
import datetime 
import ast  # For parsing stringified lists
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def parse_csv(category: Literal['web', 'pwn', 'crypto', 'reverse', 'forensics'] = "web", file_path: str = "") -> list:
    path_check = Path(file_path)
    details = []

    if path_check.exists() and path_check.is_file():
        with open(file_path, 'r') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                print(f"Processing row: {row}")  # Debug: Print each row
                if "Challenge Name" in row:
                    continue
                time_str = row[-1]
                # Parse category from stringified list, e.g., "['web']" -> "web"
                try:
                    category_list = ast.literal_eval(row[1])
                    row_category = category_list[0] if isinstance(category_list, list) and category_list else ""
                except (ValueError, SyntaxError):
                    print(f"Skipping row with invalid category format: {row[1]}")  # Debug
                    continue
                if category == row_category:
                    if time_str != "Nil":
                        try:
                            datetime.datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S.%f')  # Validate date
                            details.append(row)
                            print(f"Added solved challenge: {row}")  # Debug
                        except ValueError:
                            print(f"Skipping row with bad date format: {row}")  # Debug
                            continue
                    else:
                        details.append(row)
                        print(f"Added unsolved challenge: {row}")  # Debug
                else:
                    print(f"Skipping row: {row} (category '{row_category}' != '{category}')")  # Debug
    else:
        raise Exception(f"Invalid file path. Did you give /bin/sudo? Never mind, it was '{file_path}'")
    
    if len(details) > 0:
        print(f"Parsed {len(details)} challenges for category '{category}'")  # Debug: Total parsed
        return details
    else:
        raise Exception(f"The CSV contains no data? Why did you give me an empty file? T_T")

def sort_challs(data: List) -> List:
    """Groups and sorts by CTF -> then challenges by solve time or name, then CTFs by earliest solve date."""
    result = {}

    for details in data:
        parts = details[0].split("-")
        ctf_name = parts[0]
        chall_name = "-".join(parts[1:])
        time_str = details[-1]
        solved_by = details[2:-1]

        entry = {
            "name": chall_name,
            "solved_by": solved_by,
            "time": None if time_str == "Nil" else time_str,
        }

        if ctf_name not in result:
            result[ctf_name] = {
                "solved": [],
                "unsolved": [],
                "first_date": None
            }

        if entry["time"]:
            parsed_time = datetime.datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S.%f")
            entry["parsed_time"] = parsed_time
            result[ctf_name]["solved"].append(entry)
            if result[ctf_name]["first_date"] is None or parsed_time < result[ctf_name]["first_date"]:
                result[ctf_name]["first_date"] = parsed_time
        else:
            result[ctf_name]["unsolved"].append(entry)

    # Sort solved by time, unsolved by name
    for ctf in result:
        result[ctf]["solved"].sort(key=lambda x: x["parsed_time"])
        result[ctf]["unsolved"].sort(key=lambda x: x["name"])
        print(f"CTF {ctf}: {len(result[ctf]['solved'])} solved, {len(result[ctf]['unsolved'])} unsolved")  # Debug: Per CTF counts

    # Sort CTFs by first_date (use max date for CTFs with only unsolved challenges)
    sorted_ctfs = sorted(result.items(), key=lambda x: x[1]["first_date"] or datetime.datetime.max)
    print(f"Sorted CTFs: {[ctf[0] for ctf in sorted_ctfs]}")  # Debug: Sorted CTF order

    return sorted_ctfs

def append_ctf_data_to_excel(workbook_path: str, data: List):
    workbook_file = Path(workbook_path)
    if workbook_file.exists():
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CTF Solves"

    # Write header
    sheet.append(["CTF Name", "Date", "Challenge Name", "", "Solved By"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    current_row = 2

    for ctf_name, challenge_data in data:
        solved_challs = challenge_data["solved"]
        unsolved_challs = challenge_data["unsolved"]
        print(f"Appending for CTF {ctf_name}: {len(solved_challs)} solved, {len(unsolved_challs)} unsolved")  # Debug

        if not solved_challs and not unsolved_challs:
            print(f"Skipping CTF {ctf_name}: No challenges")  # Debug
            continue

        first_date = challenge_data["first_date"]
        formatted_date = ""
        if first_date:
            day = first_date.day
            suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
            formatted_date = first_date.strftime(f"%#d{suffix} %B")
            print(f"Formatted date for {ctf_name}: {formatted_date}")  # Debug

        start_row = current_row

        # Append solved challenges
        for challenge in solved_challs:
            sheet.append([
                "",  # CTF name will be filled via merge
                "",  # Date will be filled via merge
                challenge["name"],
                "",
                ', '.join(challenge["solved_by"])
            ])
            # Apply green fill to challenge name and solved by cells
            chall_cell = sheet.cell(row=current_row, column=3)
            chall_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            solved_by_cell = sheet.cell(row=current_row, column=5)
            solved_by_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            print(f"Appended solved challenge: {challenge['name']} at row {current_row}")  # Debug
            current_row += 1

        # Append unsolved challenges
        for challenge in unsolved_challs:
            sheet.append([
                "",  # CTF name will be filled via merge
                "",  # Date will be filled via merge
                challenge["name"],
                "",
                ', '.join(challenge["solved_by"])
            ])
            # Apply red fill to challenge name and solved by cells
            chall_cell = sheet.cell(row=current_row, column=3)
            chall_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            gap_cell = sheet.cell(row=current_row, column=4)
            gap_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            solved_by_cell = sheet.cell(row=current_row, column=5)
            solved_by_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            print(f"Appended unsolved challenge: {challenge['name']} at row {current_row}")  # Debug
            current_row += 1

        # Merge "CTF Name" and "Date" cells
        if current_row > start_row:  # Only merge if challenges were added
            ctf_range = f"A{start_row}:A{current_row - 1}"
            date_range = f"B{start_row}:B{current_row - 1}"
            sheet.merge_cells(ctf_range)
            sheet.merge_cells(date_range)

            ctf_cell = sheet.cell(row=start_row, column=1)
            ctf_cell.value = ctf_name
            ctf_cell.font = Font(bold=True)
            ctf_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            date_cell = sheet.cell(row=start_row, column=2)
            date_cell.value = formatted_date
            date_cell.font = Font(italic=True)
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            print(f"No cells merged for {ctf_name}: No challenges appended")  # Debug

    # Autofit column widths
    min_widths = [30, 20, 25, 5, 30]
    for i in range(1, 6):
        max_length = 0
        column = [sheet.cell(row=j, column=i).value for j in range(1, sheet.max_row + 1)]
        for value in column:
            if value is not None:
                max_length = max(max_length, len(str(value)))
        adjusted_width = max(max_length + 2, min_widths[i - 1])
        sheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    workbook.save(workbook_path)
    print(f"✅ Final format saved to {workbook_path}")

# Example usage
data = sort_challs(parse_csv(file_path="damctf25.csv"))
append_ctf_data_to_excel("websheet.xlsx", data)
